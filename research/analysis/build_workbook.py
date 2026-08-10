"""
Builds research/cubesat_failure_database.xlsx from the primary-source research.

Inputs (both produced by scripts in this directory, both traceable to NASA sources):
  - jacklin_appendix_a.csv   (198 missions, NASA/TM-2018-220034 Appendix A)
  - detailed_cases.py        (8 deep-dive cases read directly from primary sources)

Run: python research/analysis/build_workbook.py
"""

import csv
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent))
from detailed_cases import CASES, JACKLIN_CLASSIFICATION, SOURCES  # noqa: E402

HERE = Path(__file__).parent
OUT = HERE.parent / "cubesat_failure_database.xlsx"
JACKLIN_CSV = HERE / "jacklin_appendix_a.csv"

HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(color="FFFFFF", bold=True, size=10)
TITLE_FONT = Font(bold=True, size=13)
NOTE_FONT = Font(italic=True, size=9, color="555555")


def style_sheet(ws, header_row=1, widths=None, wrap_from=None):
    """Freeze headers, add autofilter, set widths, wrap long text columns."""
    for cell in ws[header_row]:
        if cell.value:
            cell.fill = HDR_FILL
            cell.font = HDR_FONT
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[header_row].height = 30
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    last_col = get_column_letter(ws.max_column)
    ws.auto_filter.ref = f"A{header_row}:{last_col}{ws.max_row}"
    if widths:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
    if wrap_from:
        for row in ws.iter_rows(min_row=header_row + 1, min_col=wrap_from):
            for c in row:
                c.alignment = Alignment(wrap_text=True, vertical="top")


def sheet_summary(wb):
    ws = wb.create_sheet("Summary", 0)
    ws["A1"] = "CubeSat / SmallSat Failure Research Database"
    ws["A1"].font = TITLE_FONT
    rows = [
        ("", ""),
        ("SCOPE AND HONESTY STATEMENT", ""),
        ("", "This workbook contains ONLY missions and figures traceable to primary sources that were"),
        ("", "actually retrieved and read. It is deliberately smaller than the original research brief"),
        ("", "requested. A planned multi-agent research fan-out (Dellingr, ASTERIA, OPS-SAT, university"),
        ("", "postmortems, ML-diagnosis architecture survey) FAILED to run - account spend limit - and"),
        ("", "nothing from it is represented here. Gaps are listed in the 'Research Gaps' sheet rather"),
        ("", "than filled with plausible-sounding content."),
        ("", ""),
        ("HEADLINE FIGURES", ""),
        ("Small satellites launched 2000-2016 that failed or partially failed", "41.3%"),
        ("  of which total mission failures", "24.2%"),
        ("  of which partial mission failures", "11.0%"),
        ("  of which launch vehicle failures", "6.1%"),
        ("Source", "Jacklin, NASA/TM-2018-220034"),
        ("", ""),
        ("CubeSat reliability immediately after ejection (DOA effect), 95% CI", "87.09% - 75.62%"),
        ("CubeSat reliability at 100 days, 95% CI", "73.24% - 58.94%"),
        ("CubeSat reliability at 2 years, 95% CI", "65.49% - 48.49%"),
        ("EPS share of failures after 30 days", ">40%"),
        ("COM share of failures after 90 days", "~30%"),
        ("ADCS + Payload + Structure combined", "<10%"),
        ("Source", "Langer & Bouwmeester, SSC16-X-2 (178 CubeSats)"),
        ("", ""),
        ("THE CEILING ON AUTONOMY (our analysis of NASA/TM-2018-220034 Appendix A)", ""),
        ("Missions analysed", 198),
    ]
    for k, v in JACKLIN_CLASSIFICATION["counts"].items():
        rows.append((f"  {k}", f"{v}  ({v / 198 * 100:.1f}%)"))
    rows += [
        ("", ""),
        ("KEY INTERPRETATION", ""),
        ("", "63% of NASA-listed failed missions have NO identifiable technical cause in the source."),
        ("", "16% were never heard from at all - no telemetry ever existed, so no onboard autonomy"),
        ("", "could have detected, diagnosed or acted on anything."),
        ("", "18% are recorded as failed only because no papers were published afterwards, which is"),
        ("", "weak evidence of failure at all."),
        ("", "These three facts bound how much of the failure record any autonomy could ever address."),
        ("", ""),
        ("Detailed case studies in this workbook", len(CASES)),
        ("Primary sources cited", len(SOURCES)),
    ]
    for r in rows:
        ws.append(list(r))
    ws.column_dimensions["A"].width = 62
    ws.column_dimensions["B"].width = 95
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows():
        for c in row:
            if c.value in ("SCOPE AND HONESTY STATEMENT", "HEADLINE FIGURES",
                           "THE CEILING ON AUTONOMY (our analysis of NASA/TM-2018-220034 Appendix A)",
                           "KEY INTERPRETATION"):
                c.font = Font(bold=True, size=11, color="1F3864")
    return ws


def sheet_mission_database(wb):
    ws = wb.create_sheet("Mission Database")
    if not JACKLIN_CSV.exists():
        ws["A1"] = "jacklin_appendix_a.csv not found - run extract_jacklin_entries.py first"
        return ws
    with JACKLIN_CSV.open(encoding="utf-8") as f:
        rows = list(csv.DictReader(f))
    headers = ["year", "mission", "organisation", "mass", "failure_severity",
               "never_heard_from", "failure_inferred_only_from_absence_of_publications",
               "description_verbatim", "source", "source_type"]
    ws.append([h.replace("_", " ").title() for h in headers])
    for r in rows:
        ws.append([r.get(h, "") for h in headers])
    style_sheet(ws, widths=[8, 30, 34, 10, 20, 12, 18, 90, 42, 26], wrap_from=8)
    return ws


def sheet_detailed_cases(wb):
    ws = wb.create_sheet("Detailed Cases")
    fields = [
        ("mission", 30), ("operator", 26), ("country", 12), ("launch_year", 10), ("size", 12),
        ("final_status", 34), ("primary_subsystem", 24), ("secondary_subsystems", 22),
        ("symptom", 46), ("source_text", 70), ("root_cause", 34), ("root_cause_confidence", 26),
        ("sudden_or_gradual", 18), ("precursor", 46), ("fdir_behaviour", 50), ("fdir_worked", 30),
        ("recovery_attempted", 30), ("recovery_outcome", 30), ("what_ended_mission", 26),
        ("opportunity_class", 34), ("opportunity_reasoning", 90), ("would_ml_have_helped", 70),
        ("source_quality", 22), ("confidence_in_analysis", 34),
    ]
    ws.append([f.replace("_", " ").title() for f, _ in fields])
    for c in CASES:
        ws.append([c.get(f, "") for f, _ in fields])
    style_sheet(ws, widths=[w for _, w in fields], wrap_from=6)
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 118
    return ws


def sheet_failure_modes(wb):
    ws = wb.create_sheet("Failure Modes")
    ws.append(["Category", "Count", "% of 198", "Notes"])
    for k, v in sorted(JACKLIN_CLASSIFICATION["counts"].items(), key=lambda kv: -kv[1]):
        ws.append([k, v, round(v / 198 * 100, 1), ""])
    ws.append([])
    ws.append(["METHOD", "", "", JACKLIN_CLASSIFICATION["note"]])
    ws.append(["REPRODUCE", "", "", "python research/analysis/jacklin_appendix_analysis.py"])
    style_sheet(ws, widths=[58, 10, 12, 96], wrap_from=4)
    return ws


def sheet_recovery_attempts(wb):
    ws = wb.create_sheet("Recovery Attempts")
    headers = ["Mission", "What Was Detected", "Detected By", "Autonomous or Ground",
               "Action Attempted", "Did It Work", "What Happened Next", "Outcome", "Source Quality"]
    ws.append(headers)
    for c in CASES:
        if c["recovery_outcome"].startswith("Unknown"):
            continue
        detected_by = "Ground" if "ground" in (c["recovery_attempted"] or "").lower() else "Unknown / not stated"
        if "accident" in (c["recovery_attempted"] or "").lower():
            detected_by = "Neither - incidental"
        ws.append([
            c["mission"], c["symptom"], detected_by,
            c["recovery_attempted"], c["recovery_attempted"],
            c["recovery_outcome"], c["what_ended_mission"],
            c["final_status"], c["source_quality"],
        ])
    style_sheet(ws, widths=[30, 52, 22, 34, 40, 32, 32, 40, 24], wrap_from=2)
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 76
    return ws


def sheet_fdir(wb):
    ws = wb.create_sheet("FDIR Analysis")
    headers = ["Mission", "FDIR Behaviour Observed", "Did FDIR Work",
               "Was The Fault Outside Its Fault Model", "Did Recovery Depend On The Failed Subsystem",
               "Would ML Have Helped", "Source Quality"]
    ws.append(headers)
    for c in CASES:
        outside = "Unknown / not publicly determined"
        depends = "Unknown / not publicly determined"
        if c["mission"].startswith("CSSWE"):
            outside = ("INFERENCE: yes - no rule existed for 'radio unreachable for an extended period', "
                       "so the condition was outside whatever fault model was implemented.")
            depends = ("Yes - critically. The radio was both the failed element and the only means of "
                       "reporting the failure or receiving a fix, so ground recovery was impossible by "
                       "construction. This is the strongest argument in the dataset for autonomous action.")
        if c["mission"].startswith("Delfi"):
            outside = ("INFERENCE: partly - protective responses fired, but on a misattributed cause "
                       "(bus data fault presented as subsystem faults).")
            depends = "Yes - the bus carrying the diagnostic data was itself the faulty element."
        if c["mission"].startswith("KySat-2"):
            depends = ("Yes - resets were attempted repeatedly but each reset re-entered the same "
                       "latch-up/drain condition; the recovery action could not break the loop.")
        ws.append([c["mission"], c["fdir_behaviour"], c["fdir_worked"], outside, depends,
                   c["would_ml_have_helped"], c["source_quality"]])
    style_sheet(ws, widths=[30, 56, 34, 62, 66, 70, 22], wrap_from=2)
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 128
    return ws


def sheet_opportunities(wb):
    ws = wb.create_sheet("Recovery Opportunities")
    ws.append(["Mission", "Opportunity Class", "Reasoning (tagged FACT / INFERENCE / HYPOTHESIS)",
               "Would ML Have Helped", "Confidence In Analysis"])
    for c in sorted(CASES, key=lambda x: x["opportunity_class"]):
        ws.append([c["mission"], c["opportunity_class"], c["opportunity_reasoning"],
                   c["would_ml_have_helped"], c["confidence_in_analysis"]])
    ws.append([])
    ws.append(["CLASS KEY", "A = clearly recoverable with existing capability | B = recoverable with a "
                             "reasonable added mechanism | C = containable / degraded operation | "
                             "D = probably unrecoverable, physical damage | E = insufficient information", "", "", ""])
    style_sheet(ws, widths=[30, 40, 104, 76, 36], wrap_from=2)
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 132
    return ws


def sheet_sources(wb):
    ws = wb.create_sheet("Sources")
    ws.append(["ID", "Citation", "URL", "Type", "Used For", "Key Figures"])
    for s in SOURCES:
        ws.append([s["id"], s["citation"], s["url"], s["type"], s["used_for"], s["key_figures"]])
    style_sheet(ws, widths=[6, 76, 62, 34, 50, 92], wrap_from=2)
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 84
    return ws


def sheet_gaps(wb):
    ws = wb.create_sheet("Research Gaps")
    ws.append(["Gap", "Why It Matters", "Status"])
    gaps = [
        ("Dellingr (NASA GSFC 6U) detailed anomaly history",
         "Highest-value single case available: multiple on-orbit anomalies, flatsat-based diagnosis, "
         "published lessons learned. Would materially strengthen the FDIR-failure analysis.",
         "NOT RESEARCHED - planned agent failed on spend limit. Paper is at USU DigitalCommons "
         "(SSC18) which returns HTTP 403 to automated fetch; try CORE / NTRS / GSFC mirror."),
        ("ASTERIA (JPL) comms-loss end of mission",
         "A comms-loss mission end is directly analogous to CSSWE; would test whether the "
         "power-cycle-on-timeout finding generalises.", "NOT RESEARCHED"),
        ("OPS-SAT (ESA) onboard autonomy results",
         "The only flown CubeSat explicitly built to experiment with onboard autonomy. Directly "
         "relevant to whether ML#2 is justified.", "NOT RESEARCHED"),
        ("University mission postmortems (Delfi-n3Xt, ESTCube-1, AAUSAT, MOVE-II, etc.)",
         "University teams publish unusually candid failure accounts; needed to test whether the "
         "NASA-list patterns hold in the university-class population.", "NOT RESEARCHED"),
        ("ML#2 diagnosis architecture comparison (model-based diagnosis, Bayesian nets, "
         "fault trees, case-based retrieval, RL critique)",
         "Part 7 of the brief cannot be answered rigorously without it. The current report gives a "
         "provisional position only.", "NOT RESEARCHED - provisional reasoning only"),
        ("Spacecraft telemetry anomaly benchmark datasets and their published critiques "
         "(e.g. NASA SMAP/MSL)",
         "Part 8 (training data) depends on knowing what public data exists and how trustworthy it is.",
         "NOT RESEARCHED"),
        ("Partial vs total failure split per mission in the 198-row table",
         "Appendix A is a two-column table; PDF extraction flattens the columns, so severity is "
         "not recoverable per row without re-reading the original layout.",
         "KNOWN LIMITATION - marked Unknown per row rather than guessed"),
    ]
    for g in gaps:
        ws.append(list(g))
    style_sheet(ws, widths=[52, 84, 76], wrap_from=1)
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 92
    return ws


def main():
    wb = Workbook()
    wb.remove(wb.active)
    sheet_summary(wb)
    sheet_mission_database(wb)
    sheet_detailed_cases(wb)
    sheet_failure_modes(wb)
    sheet_recovery_attempts(wb)
    sheet_fdir(wb)
    sheet_opportunities(wb)
    sheet_sources(wb)
    sheet_gaps(wb)
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"wrote {OUT}")
    print("sheets:", ", ".join(wb.sheetnames))
    return 0


if __name__ == "__main__":
    sys.exit(main())
